"""
Compliance Report Export.

Generates downloadable Excel (.xlsx) and CSV reports from live MongoDB data.

Available reports
─────────────────
  GET /export/events.csv              — raw event dump (CSV, backward-compat)
  GET /export/reports/shadow-it       — unsanctioned / unknown SaaS activity
  GET /export/reports/genai           — GenAI usage and DLP exfiltration events
  GET /export/reports/quarantine-log  — quarantine actions + firewall status
  GET /export/reports/risk-trends     — daily risk aggregation (trend data)

All Excel reports share a common style: frozen header row, auto-column widths,
conditional fill on risk level, and a summary sheet.
"""

import csv
import io
import logging
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.core.auth import get_current_user
from app.core.database import get_database

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Openpyxl helpers ──────────────────────────────────────────────────────────

def _new_workbook(sheet_name: str = "Report"):
    """Return (wb, ws) with a styled header-ready worksheet."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A2"
    return wb, ws


def _style_header(ws, headers: list) -> None:
    """Write bold, dark-background header row."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center

    ws.row_dimensions[1].height = 22


def _auto_widths(ws) -> None:
    """Set column widths based on max content length (capped at 50)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=8,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)


def _risk_fill(risk_score):
    """Return an openpyxl PatternFill colour based on risk level."""
    from openpyxl.styles import PatternFill
    try:
        score = float(risk_score or 0)
    except (ValueError, TypeError):
        score = 0
    if score >= 70:
        color = "FCE4E4"   # light red
    elif score >= 40:
        color = "FFF3CD"   # light amber
    else:
        color = "E8F5E9"   # light green
    return PatternFill(fill_type="solid", fgColor=color)


def _wb_to_stream(wb) -> io.BytesIO:
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _xlsx_response(stream: io.BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([stream.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── CSV export (backward-compatible) ─────────────────────────────────────────

@router.get("/export/events.csv")
async def export_events_csv(
    days: int = Query(7, ge=1, le=365),
    _: dict = Depends(get_current_user),
    db = Depends(get_database),
):
    """Export raw events as CSV (last N days)."""
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()
    events = await db.events.find(
        {"timestamp": {"$gte": cutoff}}
    ).sort("timestamp", -1).to_list(length=50_000)

    fields = [
        "timestamp", "source_ip", "device_name", "mac_address",
        "app_name", "bytes_sent", "bytes_received",
        "risk_score", "risk_level", "is_anomalous", "is_genai_exfiltration",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for e in events:
        writer.writerow({k: e.get(k, "") for k in fields})

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=events_{days}d.csv"},
    )


# ── Shadow IT Report ──────────────────────────────────────────────────────────

@router.get("/export/reports/shadow-it")
async def report_shadow_it(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(get_current_user),
    db = Depends(get_database),
):
    """
    Excel report: unsanctioned and unreviewed SaaS applications detected
    in the last N days with per-app risk aggregation and device breakdown.
    """
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    # Events from non-sanctioned apps
    app_profiles = await db.app_profiles.find(
        {"is_sanctioned": {"$ne": True}}
    ).to_list(length=None)
    shadow_app_names = {p["name"] for p in app_profiles}

    events = await db.events.find(
        {
            "timestamp":  {"$gte": cutoff},
            "app_name":   {"$in": list(shadow_app_names)} if shadow_app_names else {"$exists": True},
        }
    ).sort("risk_score", -1).to_list(length=50_000)

    wb, ws = _new_workbook("Shadow IT Events")
    headers = [
        "Timestamp", "Source IP", "Device Name", "MAC Address",
        "Application", "Category", "Sanctioned?",
        "Bytes Sent", "Risk Score", "Risk Level", "Anomalous?",
        "Risk Factors",
    ]
    _style_header(ws, headers)

    # Build app category lookup
    prof_map = {p["name"]: p for p in app_profiles}

    for row_idx, e in enumerate(events, start=2):
        app   = e.get("app_name", "Unknown")
        prof  = prof_map.get(app, {})
        sanctioned = prof.get("is_sanctioned")
        sanc_label = "No" if sanctioned is False else ("Yes" if sanctioned else "Unknown")

        row = [
            e.get("timestamp", ""),
            e.get("source_ip", ""),
            e.get("device_name", ""),
            e.get("mac_address", ""),
            app,
            prof.get("category", "Unknown"),
            sanc_label,
            e.get("bytes_sent", 0),
            e.get("risk_score", 0),
            e.get("risk_level", ""),
            "Yes" if e.get("is_anomalous") else "No",
            "; ".join(e.get("risk_reasons", [])),
        ]
        fill = _risk_fill(e.get("risk_score", 0))
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 9:   # Risk Score column
                cell.fill = fill

    _auto_widths(ws)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    _style_header(ws2, ["Application", "Event Count", "Avg Risk", "Max Risk", "Category"])
    app_stats: dict = {}
    for e in events:
        a = e.get("app_name", "Unknown")
        if a not in app_stats:
            app_stats[a] = {"count": 0, "risk_sum": 0.0, "max_risk": 0.0}
        app_stats[a]["count"] += 1
        r = float(e.get("risk_score", 0))
        app_stats[a]["risk_sum"] += r
        app_stats[a]["max_risk"] = max(app_stats[a]["max_risk"], r)

    for r_idx, (app, s) in enumerate(
        sorted(app_stats.items(), key=lambda x: -x[1]["count"]), start=2
    ):
        prof = prof_map.get(app, {})
        ws2.cell(row=r_idx, column=1, value=app)
        ws2.cell(row=r_idx, column=2, value=s["count"])
        ws2.cell(row=r_idx, column=3, value=round(s["risk_sum"] / s["count"], 1))
        ws2.cell(row=r_idx, column=4, value=round(s["max_risk"], 1))
        ws2.cell(row=r_idx, column=5, value=prof.get("category", "Unknown"))
    _auto_widths(ws2)

    fname = f"shadow_it_report_{days}d_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(_wb_to_stream(wb), fname)


# ── GenAI Report ──────────────────────────────────────────────────────────────

@router.get("/export/reports/genai")
async def report_genai(
    days: int = Query(30, ge=1, le=365),
    _: dict = Depends(get_current_user),
    db = Depends(get_database),
):
    """
    Excel report: all GenAI application usage, flagging DLP exfiltration events.
    """
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    events = await db.events.find(
        {
            "timestamp":  {"$gte": cutoff},
            "$or": [
                {"app_name": {"$regex": "genai|chatgpt|claude|gemini|gpt|mistral|perplexity", "$options": "i"}},
                {"is_genai_exfiltration": True},
            ],
        }
    ).sort("timestamp", -1).to_list(length=20_000)

    wb, ws = _new_workbook("GenAI Activity")
    headers = [
        "Timestamp", "Source IP", "Device Name",
        "Application", "Bytes Sent (KB)", "Risk Score",
        "DLP Exfiltration?", "Risk Factors",
    ]
    _style_header(ws, headers)

    for row_idx, e in enumerate(events, start=2):
        is_exfil = bool(e.get("is_genai_exfiltration"))
        row = [
            e.get("timestamp", ""),
            e.get("source_ip", ""),
            e.get("device_name", ""),
            e.get("app_name", ""),
            round(e.get("bytes_sent", 0) / 1024, 2),
            e.get("risk_score", 0),
            "YES ⚠" if is_exfil else "No",
            "; ".join(e.get("risk_reasons", [])),
        ]
        fill = _risk_fill(e.get("risk_score", 0))
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 6:
                cell.fill = fill

    _auto_widths(ws)

    # DLP summary sheet
    ws2 = wb.create_sheet("DLP Summary")
    _style_header(ws2, ["Metric", "Count"])
    total      = len(events)
    exfil      = sum(1 for e in events if e.get("is_genai_exfiltration"))
    high_risk  = sum(1 for e in events if e.get("risk_score", 0) >= 70)
    unique_ips = len({e.get("source_ip") for e in events})
    for r_idx, (label, val) in enumerate(
        [
            ("Total GenAI Events", total),
            ("DLP Exfiltration Events", exfil),
            ("High-Risk Events (≥70)", high_risk),
            ("Unique Source IPs", unique_ips),
            ("Report Period (days)", days),
        ],
        start=2,
    ):
        ws2.cell(row=r_idx, column=1, value=label)
        ws2.cell(row=r_idx, column=2, value=val)
    _auto_widths(ws2)

    fname = f"genai_report_{days}d_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(_wb_to_stream(wb), fname)


# ── Quarantine Log Report ─────────────────────────────────────────────────────

@router.get("/export/reports/quarantine-log")
async def report_quarantine_log(
    _: dict = Depends(get_current_user),
    db = Depends(get_database),
):
    """
    Excel report: full quarantine log with firewall enforcement status and
    the audit trail of the admin / policy that triggered each action.
    """
    quarantined = await db.quarantined_ips.find({}).sort("timestamp", -1).to_list(length=None)

    wb, ws = _new_workbook("Quarantine Log")
    headers = [
        "Timestamp", "IP Address", "Quarantined By",
        "Reason", "Firewall Status", "FW Message",
    ]
    _style_header(ws, headers)

    for row_idx, q in enumerate(quarantined, start=2):
        row = [
            q.get("timestamp", ""),
            q.get("ip", ""),
            q.get("quarantined_by", ""),
            q.get("reason", ""),
            q.get("status", ""),
            q.get("fw_message", ""),
        ]
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_widths(ws)

    # Correlated audit trail sheet
    ws2 = wb.create_sheet("Audit Trail")
    _style_header(ws2, [
        "Timestamp", "Admin", "Action", "Resource", "Outcome", "Detail"
    ])
    audit_records = await db.audit_logs.find(
        {"action": {"$regex": "QUARANTINE|BLOCK|UNQUARANTINE", "$options": "i"}}
    ).sort("timestamp", -1).to_list(length=5000)

    for r_idx, rec in enumerate(audit_records, start=2):
        ws2.cell(row=r_idx, column=1, value=rec.get("timestamp", ""))
        ws2.cell(row=r_idx, column=2, value=rec.get("admin", ""))
        ws2.cell(row=r_idx, column=3, value=rec.get("action", ""))
        ws2.cell(row=r_idx, column=4, value=f"{rec.get('resource_type','')}:{rec.get('resource_id','')}")
        ws2.cell(row=r_idx, column=5, value=rec.get("outcome", ""))
        ws2.cell(row=r_idx, column=6, value=rec.get("detail", ""))
    _auto_widths(ws2)

    fname = f"quarantine_log_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(_wb_to_stream(wb), fname)


# ── Risk Trends Report ────────────────────────────────────────────────────────

@router.get("/export/reports/risk-trends")
async def report_risk_trends(
    days: int = Query(30, ge=7, le=365),
    _: dict = Depends(get_current_user),
    db = Depends(get_database),
):
    """
    Excel report: daily aggregated risk metrics for trend analysis
    (avg risk, anomaly count, critical events, top app per day).
    """
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    pipeline = [
        {"$match": {"timestamp": {"$gte": cutoff}}},
        {"$addFields": {
            "date_str": {"$substr": ["$timestamp", 0, 10]}
        }},
        {"$group": {
            "_id":           "$date_str",
            "total_events":  {"$sum": 1},
            "avg_risk":      {"$avg": "$risk_score"},
            "max_risk":      {"$max": "$risk_score"},
            "anomalies":     {"$sum": {"$cond": ["$is_anomalous", 1, 0]}},
            "critical":      {"$sum": {"$cond": [{"$gte": ["$risk_score", 70]}, 1, 0]}},
            "unique_ips":    {"$addToSet": "$source_ip"},
            "apps":          {"$push": "$app_name"},
        }},
        {"$sort": {"_id": 1}},
    ]
    daily = await db.events.aggregate(pipeline).to_list(length=None)

    wb, ws = _new_workbook("Risk Trends")
    headers = [
        "Date", "Total Events", "Avg Risk Score", "Max Risk Score",
        "Anomalies", "Critical Events (≥70)", "Unique IPs",
        "Top Application",
    ]
    _style_header(ws, headers)

    for row_idx, day in enumerate(daily, start=2):
        # Find most-used app for the day
        from collections import Counter
        top_app = (Counter(day.get("apps", [])).most_common(1) or [("", 0)])[0][0]
        avg_r = round(day.get("avg_risk", 0), 1)

        row = [
            day["_id"],
            day.get("total_events", 0),
            avg_r,
            round(day.get("max_risk", 0), 1),
            day.get("anomalies", 0),
            day.get("critical", 0),
            len(day.get("unique_ips", [])),
            top_app,
        ]
        fill = _risk_fill(avg_r)
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3:
                cell.fill = fill

    _auto_widths(ws)

    # Overall summary sheet
    ws2 = wb.create_sheet("Overall Summary")
    _style_header(ws2, ["Metric", "Value"])
    total   = sum(d.get("total_events", 0) for d in daily)
    anom    = sum(d.get("anomalies", 0)    for d in daily)
    crit    = sum(d.get("critical", 0)     for d in daily)
    avg_all = round(sum(d.get("avg_risk", 0) for d in daily) / max(len(daily), 1), 1)
    for r_idx, (label, val) in enumerate(
        [
            ("Report Period (days)", days),
            ("Total Events", total),
            ("Total Anomalies", anom),
            ("Anomaly Rate (%)", round(anom / max(total, 1) * 100, 1)),
            ("Critical Events (≥70)", crit),
            ("Overall Avg Risk Score", avg_all),
        ],
        start=2,
    ):
        ws2.cell(row=r_idx, column=1, value=label)
        ws2.cell(row=r_idx, column=2, value=val)
    _auto_widths(ws2)

    fname = f"risk_trends_{days}d_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(_wb_to_stream(wb), fname)
